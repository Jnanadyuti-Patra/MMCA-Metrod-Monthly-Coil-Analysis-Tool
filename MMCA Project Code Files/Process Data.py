import openpyxl
from datetime import datetime
from tabulate import tabulate


def read_excel_file(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    return wb


def convert_time_to_seconds(time_str):
    # Ensure time_str is a string before converting
    if isinstance(time_str, datetime):
        time_str = time_str.strftime("%m/%d/%Y %H:%M:%S")

    # Convert time in the format "Month/Date/Year Hour:Minute:Second" to seconds since epoch
    time_format = "%m/%d/%Y %H:%M:%S"
    try:
        time_obj = datetime.strptime(time_str, time_format)
        return int(time_obj.timestamp())
    except ValueError:
        return None  # Invalid time format


def filter_rows_by_time(sheet, start_time, end_time):
    # Assume the "Time" column is the first column (index 0)
    time_col_index = 0

    # Collect rows that fall within the specified time range
    filtered_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        time_value = convert_time_to_seconds(row[time_col_index])  # Convert to seconds since epoch

        if time_value is not None and start_time <= time_value <= end_time:
            filtered_rows.append(row)

    return filtered_rows


def calculate_column_averages(filtered_rows, time_col_index):
    # Calculate averages for each column except the Time column
    num_cols = len(filtered_rows[0])
    col_sums = [0] * num_cols
    col_counts = [0] * num_cols

    for row in filtered_rows:
        for i, cell_value in enumerate(row):
            if i == time_col_index:
                continue  # Skip the Time column
            if cell_value is not None and isinstance(cell_value, (int, float)):
                col_sums[i] += cell_value
                col_counts[i] += 1

    column_averages = [col_sum / col_count if col_count > 0 else None for col_sum, col_count in
                       zip(col_sums, col_counts)]
    return column_averages
def main():
    # Specify the Excel file paths
    source_file_path = "C://Users//jnana//OneDrive//Desktop//Metrod-Internship//Process_Values.xlsx"
    destination_file_path = "C://Users//jnana//OneDrive//Desktop//Metrod-Internship//PV_Summarised.xlsx"

    try:
        # Read the Excel file
        wb = read_excel_file(source_file_path)

        # Access the Process_Values sheet
        if 'Process_Values' in wb.sheetnames:
            sheet = wb['Process_Values']
        else:
            print("Error: 'Process_Values' sheet not found in the Excel file.")
            return

        # Get user inputs for start time, end time, and grade
        start_time_str = input("Enter Start Time (Month/Date/Year Hour:Minute:Second): ")
        end_time_str = input("Enter End Time (Month/Date/Year Hour:Minute:Second): ")
        grade = input("Enter Grade: ")

        # Convert user inputs to seconds since epoch
        start_time = convert_time_to_seconds(start_time_str)
        end_time = convert_time_to_seconds(end_time_str)

        if start_time is None or end_time is None:
            print("Error: Invalid time format. Please enter time in the format 'Month/Date/Year Hour:Minute:Second'.")
            return

        # Filter rows based on time range
        filtered_rows = filter_rows_by_time(sheet, start_time, end_time)

        # Print the filtered rows as a table with averages
        if filtered_rows:
            header = [cell.value for cell in sheet[1]]
            print("\nFiltered Rows:")
            print(tabulate(filtered_rows, headers=header, tablefmt="grid"))

            # Calculate column averages, excluding the Time column
            time_col_index = 0  # Time column is the first column
            column_averages = calculate_column_averages(filtered_rows, time_col_index)

            # Add the Start Time, End Time, and Grade to the averages row
            avg_row = [start_time_str, end_time_str, grade]
            for i, header_name in enumerate(header):
                if i != time_col_index:  # Skip the Time column
                    avg_row.append(column_averages[i] if column_averages[i] is not None else None)

            # Modify the header to remove the Time column and add Start Time, End Time, Grade
            modified_header = ["Start Time", "End Time", "Grade"] + [header[i] for i in range(len(header)) if
                                                                     i != time_col_index]

            print("\nColumn Averages:")
            print(tabulate([avg_row], headers=modified_header, tablefmt="grid"))
        else:
            print("\nNo rows found within the specified time range.")

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    main()


